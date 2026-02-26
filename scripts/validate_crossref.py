#!/usr/bin/env python3
"""
Cross-reference Nunki shelter data with the National Service Provider List (NSPL).
Run from project root: python3 scripts/validate_crossref.py

Uses NSPL 2019 CSV by default (no extra deps). With openpyxl installed, uses
NSPL 2024 XLSX for latest data: pip install openpyxl
"""

import csv
import io
import json
import re
import urllib.request
from pathlib import Path

NSPL_2019_CSV = "https://open.canada.ca/data/dataset/7e0189e3-8595-4e62-a4e9-4fed6f265e10/resource/034882c4-60e0-4259-8a13-bcf9c0a93391/download/final_chpdopendatanspl_dataset-2019_june7_2020.csv"
NSPL_2024_XLSX = "https://open.canada.ca/data/dataset/7e0189e3-8595-4e62-a4e9-4fed6f265e10/resource/bfbc387f-ac55-4bf6-8656-5a14078386be/download/nspl2024_opengov_list_jun12.xlsx"


# Known name aliases: Nunki name -> NSPL name (or vice versa). Used to treat as same shelter.
SHELTER_NAME_ALIASES: dict[str, set[str]] = {
    "vancouver": {
        "aboriginal shelter", "aboriginal central street shelter",
        "lookout downtown", "lookout - downtown", "lookout downtown shelter",
        "new fountain shelter", "phs community services society - new fountain shelter",
        "lookout - al mitchell place", "lookout - al mitchell place shelter",
        "lookout - sakura-so", "lookout - walton hotel", "walton hotel",
        "lookout - hazelton", "hazelton",
    },
    "toronto": {
        "na-me-res", "na-me-res (native men's residence)", "native men's residence",
        "covenant house - madison", "covenant house - toronto", "covenant house toronto",
        "covenant house - rights of passage",
        "tsss scarborough village residence", "scarborough village residence",
        "fife house denison ave", "fife house - denison", "fife house denison",
        "fife house sherbourne st", "fife house - sherbourne",
        "christie ossington", "christie-ossington centre male", "christie ossington men's hostel south",
        "dixon hall heyworth house", "dixon hall", "dixon hall schoolhouse",
        "fred victor centre bethlehem united shelter", "fred victor",
        "street haven", "street haven at the crossroads",
        "youthlink", "youthlink ",
        "native child & family services toronto", "native child and family services",
        "native child & family - spadina",
        "birkdale residence", "tsss birkdale residence",
        "st. clare's residence", "svdp st. clare's residence",
        "amelie house", "svdp amelie house", "elisa house", "svdp elisa house",
        "salvation army - new hope leslieville", "sa new hope leslieville",
        "evangeline residence", "sa evangeline residence",
        "seaton house", "seaton house - main site", "tsss seaton house",
        "family residence - main", "tsss family residence",
    },
}


def normalize_name(s: str) -> str:
    """Lowercase, collapse whitespace, remove common suffixes for matching."""
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s.lower().strip())
    for suffix in [" shelter", " - emergency", " emergency", " (emergency)"]:
        s = s.removesuffix(suffix)
    return s.strip()


def names_match(nunki_name: str, nspl_name: str, city_id: str) -> bool:
    """Check if two shelter names refer to the same place (exact, fuzzy, or alias)."""
    nn, np = normalize_name(nunki_name), normalize_name(nspl_name)
    if nn == np:
        return True
    if similarity(nunki_name, nspl_name) >= 0.85:
        return True
    aliases = SHELTER_NAME_ALIASES.get(city_id, set())
    return nn in aliases and np in aliases


def similarity(a: str, b: str) -> float:
    """Simple word overlap. 1.0 = exact match after normalize."""
    na, nb = normalize_name(a), normalize_name(b)
    if na == nb:
        return 1.0
    wa, wb = set(na.split()), set(nb.split())
    if not wa:
        return 0.0
    return len(wa & wb) / len(wa)


def fetch_nspl_2019() -> list[dict]:
    with urllib.request.urlopen(NSPL_2019_CSV, timeout=60) as r:
        text = r.read().decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def fetch_nspl_2024() -> list[dict] | None:
    """Requires openpyxl. Returns None if not available."""
    try:
        import openpyxl
    except ImportError:
        return None
    with urllib.request.urlopen(NSPL_2024_XLSX, timeout=60) as r:
        wb = openpyxl.load_workbook(io.BytesIO(r.read()), read_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    name_col = next((i for i, h in enumerate(headers) if "shelter" in h.lower() and "nom" in h.lower() or "name" in h.lower()), 5)
    prov_col = next((i for i, h in enumerate(headers) if "province" in h.lower() and "code" in h.lower()), 1)
    city_col = next((i for i, h in enumerate(headers) if "city" in h.lower() or "ville" in h.lower()), 2)
    rows = []
    for row in ws.iter_rows(min_row=2):
        vals = [str(c.value or "").strip() if c.value is not None else "" for c in row]
        if len(vals) > max(name_col, prov_col, city_col) and vals[name_col]:
            rows.append({
                "Shelter Name/Nom du refuge": vals[name_col],
                "Province Code": vals[prov_col] if prov_col < len(vals) else "",
                "City/Ville": vals[city_col] if city_col < len(vals) else "",
            })
    wb.close()
    return rows


def fetch_nspl() -> list[dict]:
    nspl = fetch_nspl_2024()
    if nspl:
        return nspl
    return fetch_nspl_2019()


def nspl_shelters_by_city(nspl: list[dict]) -> dict[str, list[dict]]:
    """Group NSPL by city. Province: BC, ON."""
    by_city: dict[str, list[dict]] = {}
    for r in nspl:
        prov = r.get("Province Code", "") or r.get("Code de la province", "")
        city = (r.get("City/Ville") or "").strip()
        name = (r.get("Shelter Name/Nom du refuge") or "").strip()
        if not name:
            continue
        if prov == "BC":
            key = "vancouver" if "vancouver" in city.lower() else city.lower()
        elif prov == "ON":
            key = "toronto" if "toronto" in city.lower() else city.lower()
        else:
            continue
        if key not in by_city:
            by_city[key] = []
        by_city[key].append({"name": name, "city": city, "province": prov})
    return by_city


def load_nunki_shelters(city_id: str, root: Path) -> list[dict]:
    path = root / "public" / "data" / f"{city_id}.json"
    if not path.exists():
        return []
    data = json.loads(path.read_text())
    return [a for a in data.get("amenities", []) if a.get("type") == "shelter"]


def main():
    root = Path(__file__).resolve().parent.parent
    nspl = fetch_nspl_2024()
    if nspl:
        print("Fetching NSPL 2024...")
    else:
        print("Fetching NSPL 2019 (install openpyxl for 2024)...")
        nspl = fetch_nspl_2019()
    by_city = nspl_shelters_by_city(nspl)
    print(f"NSPL: {len(nspl)} shelters total. Vancouver: {len(by_city.get('vancouver', []))}, Toronto: {len(by_city.get('toronto', []))}")

    for city_id, label in [("vancouver", "Vancouver"), ("toronto", "Toronto")]:
        print(f"\n--- {label} ---")
        ours = load_nunki_shelters(city_id, root)
        nspl_list = by_city.get(city_id, [])
        nspl_names = {normalize_name(r["name"]) for r in nspl_list}

        # Ours not in NSPL (use alias-aware matching)
        missing_in_nspl = []
        for a in ours:
            n = a.get("name", "")
            matched = any(names_match(n, r["name"], city_id) for r in nspl_list)
            if matched:
                continue
            best = max(nspl_list, key=lambda r: similarity(n, r["name"])) if nspl_list else None
            sim = similarity(n, best["name"]) if best else 0
            missing_in_nspl.append((n, best["name"] if best else "", sim))

        # NSPL not in ours (exclude if we have an alias match)
        def we_have(nspl_name: str) -> bool:
            return any(names_match(a.get("name", ""), nspl_name, city_id) for a in ours)
        missing_in_ours = [r for r in nspl_list if not we_have(r["name"])]

        if missing_in_nspl:
            print(f"  In Nunki but not in NSPL (or name mismatch): {len(missing_in_nspl)}")
            for name, best_match, sim in sorted(missing_in_nspl, key=lambda x: -x[2])[:10]:
                note = f" (closest: {best_match[:40]}...)" if best_match and sim < 0.8 else ""
                print(f"    - {name}{note}")
        else:
            print("  All Nunki shelters found in NSPL (by name match).")

        if missing_in_ours:
            print(f"  In NSPL but not in Nunki: {len(missing_in_ours)} (candidates to add)")
            for r in missing_in_ours[:8]:
                print(f"    - {r['name']} ({r['city']})")
        else:
            print("  No NSPL shelters missing from Nunki.")

    print("\nDone. Use this report to spot-check and add missing shelters.")

if __name__ == "__main__":
    main()
